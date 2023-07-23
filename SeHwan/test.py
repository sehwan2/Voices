import requests
import openpyxl

url = 'https://apis.openapi.sk.com/tmap/routes/pedestrian'
appKey = 'PpgbRQ84nYWukHJFfAjA3gFoyXrOfLGazEWmhID0'

headers = {
    'Content-Type': 'application/x-www-form-urlencoded',
    'appKey': appKey,
    'Accept-Language': 'ko'
}

data = {
    'startX' : '126.92365493654832', # 출발지의 경도
    'startY' : '37.556770374096615', # 출발지의 위도
    'angle' : '1',
    'speed' : '60',
    'endPoiId' : '334852',
    'endX' : '126.92432158129688', # 도착지의 경도
    'endY' : '37.55279861528311', # 도착지의 위도
    'passList' : '126.92774822,37.55395475',
    'reqCoordType' : 'WGS84GEO',
    'startName' : '출발',
    'endName' : '도착',
    'searchOption' : '0',
    'resCoordType' : 'WGS84GEO'
}

response = requests.post(url, headers=headers, data=data)

# 엑셀 파일 생성 & 시트 선택
wb = openpyxl.Workbook()
sheet = wb.active

# 결과 파싱
result = response.json()
# 경로에 대한 정보 추출
routes = result['features'][0]['properties']['total']['subPaths'][0]['path']['coordinates']

# 시트에 데이터 입력
for i, route in enumerate(routes):
    sheet.cell(row=i+1, column=1, value=route['lon'])
    sheet.cell(row=i+1, column=2, value=route['lat'])

# 엑셀 파일 저장
wb.save('result.xlsx')